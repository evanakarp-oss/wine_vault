#!/usr/bin/env python3
"""
parse_auction_ratings.py — landing layer for critic RATINGS embedded in
auction catalogs.

Acker 261W catalogs carry, in the `WineNote` column, a critic tasting note
plus a score tag like `(93pts BH)`, `(96-98+pts VM)`, or a bare `(95pts)`.
This script extracts those into a structured, greppable JSON under
`raw/ratings/<sale>/` — the source-of-truth landing layer. The compile pass
(`compile_auction_ratings.py`) matches them to producer pages and writes a
`## Critic Ratings` section. Landing is deterministic; matching/curation is
the compile step's job (same split as scrape/parse vs compile elsewhere).

Critic initials → publication (auction convention). Unknown initials are kept
verbatim in `critic` so nothing is silently mis-attributed.

Landing is complete-by-construction: with no --file, EVERY parseable catalog
in raw/auctions/ is landed (not just the newest), so no week is ever silently
skipped. 261W is a rolling weekly sale (unsold lots persist week to week);
that's fine — the compile step dedupes ratings by (wine, vintage, critic,
score), so overlapping lots across weeks collapse to one row per rating.

Usage:
    python scripts/parse_auction_ratings.py                       # parse ALL catalogs → dry-run summary
    python scripts/parse_auction_ratings.py --apply               # write raw/ratings/<sale>/ratings_<week>.json for all
    python scripts/parse_auction_ratings.py --file raw/auctions/Catalog_261W_30.xlsx --apply   # single catalog
    python scripts/parse_auction_ratings.py --check               # exit 1 if any catalog is unlanded (CI drift guard)
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

VAULT = Path(__file__).resolve().parent.parent
AUCTIONS = VAULT / "raw" / "auctions"
RATINGS = VAULT / "raw" / "ratings"

# Critic initials as they appear in Acker WineNotes → publication label.
CRITIC_MAP = {
    "VM": "Vinous",
    "BH": "Burghound",
    "WS": "Wine Spectator",
    "WA": "Wine Advocate",
    "JS": "James Suckling",
    "JD": "Jeb Dunnuck",
    "IWC": "International Wine Cellar",
    "WE": "Wine Enthusiast",
    "D": "Decanter",
}

# ( <score>[-<hi>][+] pts|points [INITIALS] )
SCORE_TAG = re.compile(
    r"\(\s*(\d{2,3})\s*(?:[-–]\s*(\d{1,3}))?\s*(\+?)\s*(?:pts?|points?)\.?\s*([A-Za-z]{1,4})?\s*\)",
    re.IGNORECASE,
)

COLS = ["LotNo", "Quantity", "BottleName", "Vintage", "WineName", "Designation",
        "Producer", "Levels", "Low", "High", "WineType", "RegionDescription",
        "ItemLocation", "ItemWineScore", "WineNote", "AddlLots"]


def load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    hdr = [str(x) for x in rows[0]]
    return [dict(zip(hdr, r)) for r in rows[1:]]


def sale_ident(path: Path) -> tuple[str, str]:
    """Extract (sale, week) from a catalog filename, robust to naming variants:

        Catalog_261W_30.xlsx              -> ('261W', '30')
        Catalog_261W_25.xlsx              -> ('261W', '25')
        catalog_261W_week26_2026-06.xlsx  -> ('261W', '26')      # explicit week token
        catalog_261W_2026-06.xlsx         -> ('261W', '2026-06') # no week number -> date-keyed
        catalog_265DE_2026-06.xlsx        -> ('265DE', '2026-06')

    The week is only used to key the output filename; it must be unique per
    catalog so nothing overwrites another. A date fallback keeps undated-week
    catalogs lossless instead of colliding on a bare ratings.json.
    """
    stem = path.stem
    sm = re.search(r"(\d{2,4}[A-Za-z]{1,3})", stem)          # sale token e.g. 261W / 265DE
    sale = sm.group(1).upper() if sm else stem.upper()
    wm = re.search(r"week[_-]?(\d{1,2})", stem, re.IGNORECASE)
    if wm:
        return sale, wm.group(1)
    # "_<week>" immediately after the sale token (Catalog_261W_30), but not a year
    m2 = re.search(re.escape(sale) + r"[_-](\d{1,2})(?!\d)", stem, re.IGNORECASE)
    if m2:
        return sale, m2.group(1)
    dm = re.search(r"(\d{4}-\d{2})", stem)                   # date-keyed fallback
    return sale, (dm.group(1) if dm else "")


def out_path_for(sale: str, week: str) -> Path:
    """Landing path for a (sale, week). Numeric weeks keep the historical
    `ratings_week<n>.json`; date/other keys use `ratings_<key>.json`."""
    if week.isdigit():
        name = f"ratings_week{week}.json"
    elif week:
        name = f"ratings_{week}.json"
    else:
        name = "ratings.json"
    return RATINGS / sale / name


def discover_catalogs() -> list[Path]:
    """Every auction catalog we can parse: an .xlsx openpyxl can open that has
    a WineNote column. Old .xls (e.g. Zachys) and non-catalog sheets are
    skipped with a warning rather than silently dropped."""
    out = []
    for p in sorted(AUCTIONS.glob("*.xls*")):
        if p.suffix.lower() != ".xlsx":
            print(f"  ! skipping {p.name}: not .xlsx (needs conversion — landing gap)", file=sys.stderr)
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            hdr = [str(x) for x in next(wb.worksheets[0].iter_rows(min_row=1, max_row=1, values_only=True))]
        except Exception as e:
            print(f"  ! skipping {p.name}: unreadable ({e})", file=sys.stderr)
            continue
        if "WineNote" not in hdr:
            print(f"  ! skipping {p.name}: no WineNote column (not an Acker-format catalog)", file=sys.stderr)
            continue
        out.append(p)
    return out


def clean_note(note: str) -> str:
    """Prefer the critic prose (text inside the outermost double-quotes);
    else the whole note with the score tag(s) removed. Collapse whitespace,
    drop table-hostile characters."""
    note = note.strip()
    q = re.search(r'"(.+)"', note, re.DOTALL)
    prose = q.group(1) if q else SCORE_TAG.sub("", note)
    prose = prose.replace("|", "/").replace("\n", " ")
    prose = re.sub(r"\s+", " ", prose).strip(' "')
    return prose


def score_str(lo: str, hi: str, plus: str) -> str:
    s = lo
    if hi:
        s = f"{lo}-{hi}"
    if plus:
        s += "+"
    return s


def extract(rows: list[dict], sale: str, week: str, source_file: str) -> list[dict]:
    out = []
    for d in rows:
        note = (d.get("WineNote") or "").strip()
        if not note:
            continue
        tags = list(SCORE_TAG.finditer(note))
        if not tags:
            continue
        # Primary score = the first tag; keep others in `all_scores`.
        prose = clean_note(note)
        all_scores = []
        for m in tags:
            initials = (m.group(4) or "").upper()
            all_scores.append({
                "critic_raw": initials or "",
                "critic": CRITIC_MAP.get(initials, initials) if initials else "",
                "score": score_str(m.group(1), m.group(2), m.group(3)),
            })
        primary = all_scores[0]
        lot = d.get("LotNo")
        source = f"{sale}·W{week} lot {lot}" if week else f"{sale} lot {lot}"
        out.append({
            "lot": lot,
            "source": source,
            "producer_raw": (d.get("Producer") or "").strip(),
            "wine": (d.get("WineName") or "").strip(),
            "designation": (d.get("Designation") or "").strip(),
            "vintage": str(d.get("Vintage") or "").strip(),
            "region": (d.get("RegionDescription") or "").strip(),
            "location": (d.get("ItemLocation") or "").strip(),
            "critic": primary["critic"],
            "critic_raw": primary["critic_raw"],
            "score": primary["score"],
            "all_scores": all_scores,
            "note": prose[:500],
        })
    return out


def parse_one(path: Path, apply: bool) -> dict:
    """Parse a single catalog; write JSON if apply. Returns a summary dict."""
    sale, week = sale_ident(path)
    rel = str(path.relative_to(VAULT)) if path.is_relative_to(VAULT) else str(path)
    rows = load_rows(path)
    ratings = extract(rows, sale, week, rel)
    out_path = out_path_for(sale, week)

    payload = {
        "sale": sale,
        "week": week,
        "source_file": rel,
        "captured": date.today().isoformat(),
        "count": len(ratings),
        "ratings": ratings,
    }
    if apply:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"path": path, "sale": sale, "week": week, "lots": len(rows),
            "ratings": len(ratings), "out": out_path}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", help="Single catalog xlsx to parse (default: ALL catalogs in raw/auctions/)")
    ap.add_argument("--apply", action="store_true", help="Write JSON. Default is dry-run summary.")
    ap.add_argument("--check", action="store_true",
                    help="Exit 1 if any parseable catalog has no landed ratings JSON (CI drift guard).")
    args = ap.parse_args()

    if args.file:
        path = Path(args.file)
        if not path.is_absolute():
            path = VAULT / path
        catalogs = [path]
    else:
        catalogs = discover_catalogs()
        if not catalogs:
            sys.exit("no parseable catalogs in raw/auctions/")

    # --check: report any catalog whose landing JSON is missing; change nothing.
    if args.check:
        missing = []
        for p in catalogs:
            sale, week = sale_ident(p)
            if not out_path_for(sale, week).exists():
                missing.append((p.name, sale, week))
        if missing:
            print(f"UNLANDED catalogs ({len(missing)}) — run: python scripts/parse_auction_ratings.py --apply")
            for name, sale, week in missing:
                print(f"  {name}  ->  {sale} W{week}")
            return 1
        print(f"all {len(catalogs)} catalogs landed ✓")
        return 0

    summaries = [parse_one(p, args.apply) for p in catalogs]
    print(f"{'wrote' if args.apply else '(dry-run) would write'} {len(summaries)} ratings file(s):\n")
    total = 0
    for s in sorted(summaries, key=lambda s: (s["sale"], s["week"])):
        total += s["ratings"]
        print(f"  {s['sale']:>6} W{s['week']:<8} {s['lots']:>5} lots -> {s['ratings']:>4} ratings   "
              f"{s['out'].relative_to(VAULT)}")
    print(f"\ntotal ratings across catalogs: {total}")
    if not args.apply:
        print("(dry-run — pass --apply to write)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
