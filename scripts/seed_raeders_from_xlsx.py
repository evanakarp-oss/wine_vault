"""
Parse the AI-compiled Desktop raeders.xlsx into a normalized snapshot under raw/raeders/.

Source xlsx has an unusual shape: 848 single-cell rows, each containing a CSV line with
an LLM citation suffix like ` [cite: 3, 4]`. Example row:
    Chateau Petrus,2006,750ml,Red Wine,"$4,999.99","Pomerol (WE: 95-97, WA: 96)" [cite: 3, 4]

We normalize to a parseable CSV and emit a markdown snapshot with YAML frontmatter.
Data is provenance-tagged as AI-compiled so downstream consumers know to treat it
as approximate (hallucinations possible) until validated against the live site.
"""
from __future__ import annotations

import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dep: openpyxl. Run: python -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(2)


VAULT = Path(__file__).resolve().parent.parent
SRC_XLSX = Path(r"C:/Users/Evan Karp/OneDrive/Desktop/raeders.xlsx")
OUT_DIR = VAULT / "raw" / "raeders"
OUT_CSV = OUT_DIR / f"raeders_inventory_{date.today().isoformat()}.csv"
OUT_MD = OUT_DIR / f"raeders_inventory_{date.today().isoformat()}.md"
OUT_REPORT = OUT_DIR / f"raeders_parse_report_{date.today().isoformat()}.md"

CITE_RE = re.compile(r"\s*\[cite:\s*[\d\s,]+\]\s*$")


def strip_citations(line: str) -> str:
    return CITE_RE.sub("", line).strip()


def load_rows() -> list[str]:
    wb = load_workbook(SRC_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        cell = r[0] if r else None
        if cell is None:
            continue
        text = str(cell).strip()
        if text:
            rows.append(text)
    return rows


def parse_rows(raw_rows: list[str]) -> tuple[list[dict], list[tuple[int, str, str]]]:
    """Returns (parsed, errors). errors = list of (line_index, raw_line, reason)."""
    if not raw_rows:
        return [], []

    # First line is the header
    header_raw = strip_citations(raw_rows[0])
    header = next(csv.reader([header_raw]))
    expected = ["Name", "Vintage", "Format", "Category", "Price", "Region/Notes"]
    if header != expected:
        print(f"WARN: header mismatch\n  got:      {header}\n  expected: {expected}", file=sys.stderr)

    parsed: list[dict] = []
    errors: list[tuple[int, str, str]] = []

    for i, line in enumerate(raw_rows[1:], start=2):
        stripped = strip_citations(line)
        try:
            row = next(csv.reader([stripped]))
        except Exception as e:
            errors.append((i, line, f"csv_error: {e}"))
            continue
        if len(row) != len(header):
            errors.append((i, line, f"column_count: got {len(row)}, expected {len(header)}"))
            continue
        record = dict(zip(header, row))

        # Normalize price: "$4,999.99" -> 4999.99
        price_raw = record.get("Price", "")
        price_clean = re.sub(r"[^\d.]", "", price_raw)
        try:
            record["price_usd"] = float(price_clean) if price_clean else None
        except ValueError:
            record["price_usd"] = None

        # Normalize vintage: accept YYYY or NV
        v = (record.get("Vintage") or "").strip()
        record["vintage"] = v if v else ""

        # Producer guess: first token of Name, split on " - " or "," when present.
        # Most rows have format like "Producer - Wine Name" or "Producer, Wine" or "Wine Name"
        name = (record.get("Name") or "").strip()
        producer_guess = ""
        for sep in [" - ", "—", " – ", " / "]:
            if sep in name:
                producer_guess = name.split(sep, 1)[0].strip()
                break
        # Fallback: first comma
        if not producer_guess and "," in name:
            producer_guess = name.split(",", 1)[0].strip()
        # Last resort: whole name
        if not producer_guess:
            producer_guess = name
        record["producer_guess"] = producer_guess

        parsed.append(record)

    return parsed, errors


def write_csv(rows: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "Name", "producer_guess", "Vintage", "Format", "Category",
            "Price", "price_usd", "Region/Notes",
        ])
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in w.fieldnames})


def write_markdown(rows: list[dict], errors: list) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    lines = [
        "---",
        "type: retailer_inventory_snapshot",
        "retailer: raeders",
        f"snapshot_date: \"{date.today().isoformat()}\"",
        "source: \"OneDrive/Desktop/raeders.xlsx (AI-compiled, brute-force scrape of raederswine.com)\"",
        "provenance_warning: \"Data was LLM-compiled, not fetched from the live site. Treat as approximate \"",
        "  \"until validated. Once the live Raeder's crawler runs, prefer its output over this.\"",
        f"row_count: {len(rows)}",
        f"parse_errors: {len(errors)}",
        "---",
        "",
        "# Raeder's inventory snapshot",
        "",
        f"Compiled from `{SRC_XLSX.name}` on {date.today().isoformat()}. {len(rows)} wines parsed.",
        "",
        "## Summary",
        "",
    ]

    # Category breakdown
    from collections import Counter
    cats = Counter(r.get("Category", "") for r in rows)
    lines.append("### By category")
    lines.append("")
    for cat, n in cats.most_common():
        lines.append(f"- {cat or '(blank)'}: {n}")
    lines.append("")

    # Top producers
    prods = Counter(r.get("producer_guess", "") for r in rows if r.get("producer_guess"))
    lines.append("### Top 20 producers (by # of SKUs)")
    lines.append("")
    for prod, n in prods.most_common(20):
        lines.append(f"- {prod}: {n}")
    lines.append("")

    # Price distribution
    prices = [r["price_usd"] for r in rows if r.get("price_usd") is not None]
    if prices:
        prices_sorted = sorted(prices)
        lines.append("### Price (USD)")
        lines.append("")
        lines.append(f"- min: ${prices_sorted[0]:.2f}")
        lines.append(f"- median: ${prices_sorted[len(prices_sorted)//2]:.2f}")
        lines.append(f"- max: ${prices_sorted[-1]:.2f}")
        lines.append(f"- count with price: {len(prices)}")
        lines.append("")

    lines.append("## Sample rows")
    lines.append("")
    lines.append("| Producer (guess) | Name | Vintage | Format | Price | Region/Notes |")
    lines.append("|---|---|---|---|---|---|")
    for r in rows[:30]:
        def esc(v):
            return str(v).replace("|", "\\|")
        lines.append(
            f"| {esc(r.get('producer_guess',''))} | {esc(r.get('Name',''))} "
            f"| {esc(r.get('Vintage',''))} | {esc(r.get('Format',''))} "
            f"| {esc(r.get('Price',''))} | {esc(r.get('Region/Notes',''))} |"
        )
    lines.append("")
    lines.append(f"*Showing 30 of {len(rows)}. Full data: `{OUT_CSV.name}`.*")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_report(errors: list) -> None:
    lines = [
        "---",
        "type: parse_report",
        "retailer: raeders",
        f"generated: \"{date.today().isoformat()}\"",
        f"errors: {len(errors)}",
        "---",
        "",
        "# Raeder's xlsx parse report",
        "",
    ]
    if not errors:
        lines.append("No parse errors.")
    else:
        lines.append(f"{len(errors)} row(s) failed to parse:")
        lines.append("")
        for i, raw, reason in errors[:50]:
            lines.append(f"- line {i}: {reason}")
            lines.append(f"  ```")
            lines.append(f"  {raw}")
            lines.append(f"  ```")
        if len(errors) > 50:
            lines.append(f"\n... and {len(errors) - 50} more")
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    if not SRC_XLSX.exists():
        print(f"Source not found: {SRC_XLSX}", file=sys.stderr)
        return 1

    raw_rows = load_rows()
    print(f"Read {len(raw_rows)} raw rows from {SRC_XLSX.name}")

    parsed, errors = parse_rows(raw_rows)
    print(f"Parsed {len(parsed)} records; {len(errors)} errors")

    write_csv(parsed)
    write_markdown(parsed, errors)
    write_report(errors)

    print(f"\nOK: wrote {OUT_CSV.name}, {OUT_MD.name}, {OUT_REPORT.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
